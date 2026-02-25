"""
Excel exporter — writes filtered tenders to a formatted .xlsx file.

The workbook has two sheets:
  1. "Matched Tenders"   — all tenders that passed the filter (colour-coded by score)
  2. "All Tenders"       — every tender found today, for reference

Colour scheme (fill colour in the Score column):
  80-100: Dark green  — excellent match
  60-79:  Green       — good match
  30-59:  Yellow      — possible match
  0-29:   Red/grey    — poor match (only on All Tenders sheet)
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from scrapers.models import Tender
import config

logger = logging.getLogger(__name__)

# ── Colour fills ──────────────────────────────────────────────────────────────
FILL_EXCELLENT = PatternFill("solid", fgColor="1A7A3C")   # Dark green
FILL_GOOD      = PatternFill("solid", fgColor="4CAF50")   # Green
FILL_POSSIBLE  = PatternFill("solid", fgColor="FFC107")   # Amber
FILL_POOR      = PatternFill("solid", fgColor="B0BEC5")   # Grey
FILL_HEADER    = PatternFill("solid", fgColor="1B3A6B")   # Navy blue header
FILL_ALT_ROW   = PatternFill("solid", fgColor="F0F4FF")   # Light blue alt row

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_TITLE   = Font(name="Calibri", bold=True, color="1B3A6B", size=10)
FONT_BODY    = Font(name="Calibri", size=10)
FONT_SCORE   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7E5"),
    right=Side(style="thin", color="D0D7E5"),
    top=Side(style="thin", color="D0D7E5"),
    bottom=Side(style="thin", color="D0D7E5"),
)

COLUMN_DEFS = [
    # (header,           width, attr_or_callable)
    ("#",                5,     None),
    ("Score",            8,     "match_score"),
    ("Portal",           14,    "portal"),
    ("Tender ID",        20,    "tender_id"),
    ("Title",            48,    "title"),
    ("Department",       28,    "department"),
    ("Location",         18,    "location"),
    ("Budget",           16,    "display_budget"),
    ("Budget OK?",       11,    "budget_in_range"),
    ("Published",        14,    "display_published"),
    ("Deadline",         18,    "display_deadline"),
    ("Matched Keywords", 36,    "matched_keywords"),
    ("Link",             50,    "url"),
]


def _score_fill(score: int) -> PatternFill:
    if score >= 80:
        return FILL_EXCELLENT
    if score >= 60:
        return FILL_GOOD
    if score >= 30:
        return FILL_POSSIBLE
    return FILL_POOR


def _get_value(tender: Tender, attr: str):
    """Extract a display value from a Tender using attribute name or method name."""
    if attr is None:
        return ""
    val = getattr(tender, attr, None)
    if callable(val):
        val = val()
    if isinstance(val, list):
        return ", ".join(str(k).replace("LOC:", "📍").replace("EXCLUDED:", "❌") for k in val)
    if isinstance(val, bool) or val is True or val is False:
        return "✓" if val else "✗"
    if val is None:
        return "—"
    return val


def _write_sheet(
    ws,
    tenders: List[Tender],
    title: str,
    run_date: str,
) -> None:
    """Write the tender list into a worksheet."""

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMN_DEFS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"{title}  |  Run: {run_date}  |  {len(tenders)} result(s)"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1B3A6B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (header, width, _) in enumerate(COLUMN_DEFS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, tender in enumerate(tenders, start=1):
        excel_row = row_idx + 2
        alt = (row_idx % 2 == 0)

        for col_idx, (header, _, attr) in enumerate(COLUMN_DEFS, start=1):
            col_letter = get_column_letter(col_idx)

            if header == "#":
                value = row_idx
            elif header == "Link":
                value = _get_value(tender, attr)
            else:
                value = _get_value(tender, attr)

            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.font = FONT_BODY
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(header in ("Title", "Department", "Matched Keywords")),
            )

            # Alternate row shading
            if alt and header not in ("Score",):
                cell.fill = FILL_ALT_ROW

            # Score column — colour by score value
            if header == "Score":
                cell.fill = _score_fill(tender.match_score)
                cell.font = FONT_SCORE
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Title — make it bold
            if header == "Title":
                cell.font = FONT_TITLE

            # Link — make it a hyperlink
            if header == "Link" and value and value != "—":
                cell.hyperlink = str(value)
                cell.font = Font(
                    name="Calibri", size=10, color="0563C1", underline="single"
                )
                cell.value = "Open ↗"

        # Row height — taller for wrapped cells
        ws.row_dimensions[excel_row].height = 36

    # ── Freeze panes & auto-filter ────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMN_DEFS))}{len(tenders) + 2}"


def export_to_excel(
    matched: List[Tender],
    all_tenders: List[Tender],
    output_dir: str = None,
) -> str:
    """
    Write two-sheet Excel file and return the file path.

    Args:
        matched:     Filtered tenders (Matched Tenders sheet).
        all_tenders: All scraped tenders (All Tenders sheet).
        output_dir:  Directory to save the file. Defaults to config.OUTPUT_DIR.

    Returns:
        Absolute path of the saved .xlsx file.
    """
    out_dir = Path(output_dir or config.OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    run_date = datetime.now().strftime("%d %b %Y %H:%M")
    date_tag = datetime.now().strftime("%Y-%m-%d")
    filename = config.OUTPUT_FILENAME.format(date=date_tag)
    filepath = out_dir / filename

    wb = openpyxl.Workbook()

    # ── Sheet 1: Matched Tenders ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Matched Tenders"
    _write_sheet(ws1, matched, "Matched Tenders — Your Best Opportunities", run_date)

    # ── Sheet 2: All Tenders ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Tenders (Raw)")
    _write_sheet(ws2, all_tenders, "All Scraped Tenders Today", run_date)

    wb.save(filepath)
    logger.info("Excel saved: %s", filepath.resolve())
    return str(filepath.resolve())
