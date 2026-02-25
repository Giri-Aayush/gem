"""
webapp.py — Local tender dashboard.

Run:  python webapp.py
Open: http://localhost:5000
"""

import importlib
import logging
import threading
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml
from flask import Flask, Response, jsonify, request

import config
from filters.tender_filter import filter_tenders
from output_engine.excel_exporter import export_to_excel
from scrapers.models import Tender

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s — %(message)s")
log = logging.getLogger("webapp")

app = Flask(__name__)

# ── Global state ──────────────────────────────────────────────────────────────
matched_tenders: list = []
all_tenders: list = []
scrape_status: str = "idle"
last_run_ts: str = ""
scrape_error: str = ""
scrape_progress: int = 0    # 0-100
scrape_stage: str = ""      # human-readable description of current step
_lock = threading.Lock()

# ── Serialiser ────────────────────────────────────────────────────────────────

def _days_left(dl_date) -> int | None:
    """Return days until deadline, or None if unknown. Negative = already passed."""
    if dl_date is None:
        return None
    today = datetime.now().date()
    target = dl_date.date() if isinstance(dl_date, datetime) else dl_date
    return (target - today).days


def _to_dict(t, rank: int) -> dict:
    if isinstance(t, Tender):
        kws = [k.replace("LOC:", "").replace("EXCLUDED:", "").strip() for k in t.matched_keywords]
        return {
            "rank": rank, "score": t.match_score,
            "title": t.title, "url": t.url,
            "department": t.department, "portal": t.portal,
            "deadline": t.display_deadline(), "budget": t.display_budget(),
            "keywords": ", ".join(kws),
            "location_match": bool(t.location_match),
            "budget_ok": t.budget_in_range,
            "days_left": _days_left(t.deadline),
        }
    # Already a dict (loaded from Excel at startup) — ensure score is int
    return dict(t, rank=rank, score=int(float(t.get("score", 0))))

# ── Startup: load latest Excel ────────────────────────────────────────────────

def _load_latest_excel():
    global matched_tenders, all_tenders, last_run_ts
    reports = Path(config.OUTPUT_DIR)
    files = sorted(
        [f for f in reports.glob("tenders_*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime, reverse=True,
    )
    if not files:
        log.info("No Excel files found in %s — start a scrape to populate data.", reports)
        return

    latest = files[0]
    log.info("Loading data from %s …", latest.name)
    try:
        wb = openpyxl.load_workbook(latest, data_only=True)  # read_only=False needed for hyperlinks
        matched_tenders = _read_sheet(wb, "Matched Tenders")
        all_tenders     = _read_sheet(wb, "All Tenders (Raw)")
        wb.close()
        mtime = datetime.fromtimestamp(latest.stat().st_mtime)
        last_run_ts = mtime.strftime("%d %b %Y %H:%M")
        log.info("Loaded %d matched / %d total from Excel.", len(matched_tenders), len(all_tenders))
    except Exception as exc:
        log.error("Could not read Excel: %s", exc)


def _read_sheet(wb, sheet_name: str) -> list:
    """
    Read tender rows from a worksheet.
    Column order mirrors COLUMN_DEFS in excel_exporter.py (1-indexed):
      1=#  2=Score  3=Portal  4=TenderID  5=Title  6=Dept  7=Loc
      8=Budget  9=BudgetOK  10=Published  11=Deadline  12=Keywords  13=Link
    """
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        return []

    rows = list(ws.iter_rows(min_row=3))  # skip title + header rows
    results = []
    for row in rows:
        if len(row) < 13:
            continue
        def v(col):  # 1-indexed column → cell value
            cell = row[col - 1]
            val = cell.value
            return str(val).strip() if val is not None else ""

        score_raw = v(2)
        try:
            score = int(float(score_raw))
        except (ValueError, TypeError):
            continue  # skip header/empty rows

        # URL stored as hyperlink on the "Open ↗" cell
        link_cell = row[12]  # col 13, 0-indexed = 12
        url = ""
        if link_cell.hyperlink:
            url = link_cell.hyperlink.target or ""

        deadline_str = v(11)
        days_left = None
        if deadline_str and deadline_str != "—":
            try:
                dl = datetime.strptime(deadline_str, "%d %b %Y %H:%M")
                days_left = _days_left(dl)
            except ValueError:
                pass

        results.append({
            "score":      score,
            "portal":     v(3),
            "tender_id":  v(4),
            "title":      v(5),
            "department": v(6),
            "location":   v(7),
            "budget":     v(8),
            "budget_ok":  v(9),
            "published":  v(10),
            "deadline":   deadline_str,
            "keywords":   v(12),
            "url":        url,
            "days_left":  days_left,
        })
    return results

# ── Background scrape ─────────────────────────────────────────────────────────

def _set_progress(pct: int, stage: str):
    """Thread-safe helper to update scrape progress and stage."""
    global scrape_progress, scrape_stage
    scrape_progress = pct
    scrape_stage = stage
    log.info("[%d%%] %s", pct, stage)


def _do_scrape():
    global matched_tenders, all_tenders, scrape_status, last_run_ts, scrape_error, scrape_progress, scrape_stage
    try:
        _set_progress(0, "Initialising scrapers…")

        from scrapers.gem_scraper import GemScraper
        from scrapers.cppp_scraper import CpppScraper

        portals = config.CONTRACTOR_PROFILE.get("portals", {})
        scraper_list = []
        if portals.get("gem", True):
            scraper_list.append(("GeM", GemScraper()))
        if portals.get("cppp", True):
            scraper_list.append(("CPPP", CpppScraper()))

        raw: list = []
        n = len(scraper_list)
        # GeM takes ~90% of the time; apportion progress 5% → 72% across scrapers
        for i, (label, scraper) in enumerate(scraper_list):
            start_pct = 5 + i * 67 // n
            end_pct   = 5 + (i + 1) * 67 // n
            _set_progress(start_pct, f"Connecting to {label} portal…")
            batch = scraper.run()
            raw.extend(batch)
            _set_progress(end_pct, f"{label} done — {len(batch)} tenders found")

        _set_progress(75, f"Scoring and filtering {len(raw)} tenders against your profile…")
        matched = filter_tenders(raw, config.CONTRACTOR_PROFILE, min_score=config.DEFAULT_MIN_SCORE)

        _set_progress(88, f"{len(matched)} matches found — saving Excel report…")
        try:
            export_to_excel(matched, raw)
        except Exception as exc:
            log.warning("Excel export failed (non-fatal): %s", exc)

        _set_progress(98, f"Wrapping up — {len(matched)} matched / {len(raw)} total")

        with _lock:
            all_tenders     = raw
            matched_tenders = matched
            last_run_ts     = datetime.now().strftime("%d %b %Y %H:%M")
            scrape_status   = "done"
            scrape_error    = ""
            scrape_progress = 100
            scrape_stage    = f"Complete — {len(matched)} matched, {len(raw)} total tenders"

        log.info("Scrape complete — %d matched / %d total.", len(matched), len(raw))

    except Exception as exc:
        log.error("Scrape failed: %s", exc, exc_info=True)
        with _lock:
            scrape_status   = "error"
            scrape_error    = str(exc)
            scrape_stage    = f"Error: {exc}"

# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/api/tenders")
def api_matched():
    return jsonify([_to_dict(t, i + 1) for i, t in enumerate(matched_tenders)])

@app.get("/api/all-tenders")
def api_all():
    return jsonify([_to_dict(t, i + 1) for i, t in enumerate(all_tenders)])

@app.get("/api/status")
def api_status():
    return jsonify({
        "status":   scrape_status,
        "last_run": last_run_ts,
        "error":    scrape_error,
        "progress": scrape_progress,
        "stage":    scrape_stage,
    })

@app.post("/run-scraper")
def trigger_scrape():
    global scrape_status
    with _lock:
        if scrape_status == "running":
            return jsonify({"error": "already running"}), 409
        scrape_status = "running"
    t = threading.Thread(target=_do_scrape, daemon=True)
    t.start()
    return jsonify({"status": "started"}), 202

@app.get("/api/reports")
def api_reports():
    """List all historical Excel files in reports/ folder."""
    reports = Path(config.OUTPUT_DIR)
    files = sorted(
        [f for f in reports.glob("tenders_*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime, reverse=True,
    )
    result = []
    for f in files:
        mtime = datetime.fromtimestamp(f.stat().st_mtime)
        # Quick row count from just the matched sheet (fast: open read-only)
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            ws = wb["Matched Tenders"] if "Matched Tenders" in wb.sheetnames else None
            rows = (ws.max_row - 2) if ws else 0  # subtract title + header rows
            wb.close()
        except Exception:
            rows = "?"
        result.append({
            "filename": f.name,
            "date": mtime.strftime("%d %b %Y"),
            "time": mtime.strftime("%H:%M"),
            "matched": rows,
            "size_kb": round(f.stat().st_size / 1024, 1),
        })
    return jsonify(result)

@app.get("/reports/<filename>")
def download_report(filename: str):
    """Serve an Excel file for download."""
    from flask import send_from_directory, abort
    if not filename.startswith("tenders_") or not filename.endswith(".xlsx"):
        abort(404)
    reports_dir = Path(config.OUTPUT_DIR).resolve()
    return send_from_directory(reports_dir, filename, as_attachment=True)

@app.get("/api/report-data/<filename>")
def api_report_data(filename: str):
    """Load matched + all tenders from a specific historical Excel file."""
    from flask import abort
    if not filename.startswith("tenders_") or not filename.endswith(".xlsx"):
        abort(400)
    f = Path(config.OUTPUT_DIR) / filename
    if not f.exists():
        abort(404)
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        matched = _read_sheet(wb, "Matched Tenders")
        all_t   = _read_sheet(wb, "All Tenders (Raw)")
        wb.close()
        m_dicts = [_to_dict(t, i + 1) for i, t in enumerate(matched)]
        a_dicts = [_to_dict(t, i + 1) for i, t in enumerate(all_t)]
        return jsonify({"matched": m_dicts, "all": a_dicts})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

_PROFILE_FILE = Path(__file__).parent / "my_profile.yaml"

@app.get("/api/settings")
def api_get_settings():
    """Return current my_profile.yaml settings as JSON."""
    with open(_PROFILE_FILE, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return jsonify({
        "your_name":                 data.get("your_name", ""),
        "locations":                 data.get("locations", []),
        "my_work_types":             data.get("my_work_types", []),
        "exclude_these_work_types":  data.get("exclude_these_work_types", []),
        "budget_minimum":            data.get("budget", {}).get("minimum", 100000),
        "budget_maximum":            data.get("budget", {}).get("maximum", 20000000),
        "minimum_match_score":       data.get("minimum_match_score", 30),
        "run_every_day_at":          data.get("run_every_day_at", "08:00"),
    })

@app.post("/api/settings")
def api_save_settings():
    """Write updated settings to my_profile.yaml and reload config module."""
    body = request.get_json(force=True)

    # Preserve email + look_back_days from existing file
    try:
        with open(_PROFILE_FILE, encoding="utf-8") as f:
            existing = yaml.safe_load(f) or {}
    except Exception:
        existing = {}

    new_data = {
        "your_name":                body.get("your_name", existing.get("your_name", "Contractor")),
        "locations":                body.get("locations", []),
        "my_work_types":            body.get("my_work_types", []),
        "budget": {
            "minimum": int(body.get("budget_minimum", 100000)),
            "maximum": int(body.get("budget_maximum", 20000000)),
        },
        "exclude_these_work_types": body.get("exclude_these_work_types", []),
        "run_every_day_at":         str(body.get("run_every_day_at", existing.get("run_every_day_at", "08:00"))),
        "email":                    existing.get("email", {"send_email": "no"}),
        "minimum_match_score":      int(body.get("minimum_match_score", 30)),
        "look_back_days":           existing.get("look_back_days", 1),
    }

    try:
        with open(_PROFILE_FILE, "w", encoding="utf-8") as f:
            yaml.dump(new_data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
    except Exception as exc:
        return jsonify({"error": f"Could not write profile: {exc}"}), 500

    # Reload config so next scrape picks up new values immediately
    try:
        importlib.reload(config)
    except Exception as exc:
        log.warning("Config reload failed: %s", exc)

    return jsonify({"status": "saved"})

@app.get("/")
def index():
    return Response(HTML, mimetype="text/html")

# ── HTML + CSS + JS (single-file dashboard) ───────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Tender Dashboard</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;color:#222}

/* Header */
header{background:#1b3a6b;color:#fff;padding:14px 24px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 6px rgba(0,0,0,.3)}
header h1{font-size:1.25rem;font-weight:700;letter-spacing:.5px}
.hdr-right{display:flex;align-items:center;gap:14px;font-size:.88rem}
#last-run{opacity:.8}
#run-btn{background:#f0a500;color:#000;border:none;padding:8px 18px;border-radius:6px;font-weight:700;cursor:pointer;font-size:.9rem;transition:background .2s}
#run-btn:hover{background:#ffc107}
#run-btn:disabled{background:#888;cursor:not-allowed;color:#ccc}
#spinner{display:none;color:#ffc107;font-weight:600;animation:pulse 1.2s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}

/* Tabs */
.tabs{background:#fff;border-bottom:2px solid #dde3ec;padding:0 24px;display:flex;gap:4px}
.tab{padding:12px 20px;border:none;background:none;cursor:pointer;font-size:.95rem;font-weight:600;color:#555;border-bottom:3px solid transparent;margin-bottom:-2px;transition:all .15s}
.tab.active{color:#1b3a6b;border-bottom-color:#1b3a6b}
.tab:hover:not(.active){background:#f5f7fb}

/* Controls bar */
.controls{display:flex;align-items:center;gap:10px;padding:12px 24px;background:#fff;border-bottom:1px solid #dde3ec;flex-wrap:wrap}
#search-box{padding:8px 12px;border:1px solid #c5cdd8;border-radius:6px;font-size:.9rem;width:280px}
#portal-filter{padding:8px 10px;border:1px solid #c5cdd8;border-radius:6px;font-size:.9rem}
#result-count{margin-left:auto;color:#666;font-size:.88rem;font-weight:600}

/* Table */
.table-wrap{overflow-x:auto;padding:16px 24px}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.1)}
thead tr{background:#1b3a6b;color:#fff}
th{padding:11px 12px;text-align:left;font-size:.83rem;font-weight:600;white-space:nowrap}
td{padding:10px 12px;font-size:.88rem;border-bottom:1px solid #eef1f6;vertical-align:top}
tbody tr:last-child td{border-bottom:none}
tbody tr:hover{filter:brightness(.96)}

/* Row colours by score */
tr.s-excellent{background:#d4edda}
tr.s-good{background:#eaf7eb}
tr.s-possible{background:#fffde7}
tr.s-poor{background:#f5f5f5}

/* Score badge */
.badge{display:inline-block;padding:3px 9px;border-radius:12px;font-size:.8rem;font-weight:700;color:#fff;white-space:nowrap}
.badge.excellent{background:#1a7a3c}
.badge.good{background:#4caf50}
.badge.possible{background:#ffc107;color:#333}
.badge.poor{background:#90a4ae;color:#fff}

/* Title link */
td a{color:#1b3a6b;text-decoration:none;font-weight:500}
td a:hover{text-decoration:underline}

/* Keywords cell */
.kw{font-size:.78rem;color:#555;max-width:200px}
.dept{max-width:180px;font-size:.83rem}

/* Days-left badge */
.dl-badge{display:inline-block;padding:3px 9px;border-radius:10px;font-size:.78rem;font-weight:700;white-space:nowrap}
.dl-expired{background:#e8e8e8;color:#777}
.dl-urgent{background:#fde8e8;color:#c62828}
.dl-soon{background:#fff3e0;color:#bf5000}
.dl-ok{background:#e8f5e9;color:#2e7d32}

/* No data */
.empty{text-align:center;padding:48px;color:#888;font-size:1rem}

/* Past reports */
.reports-grid{padding:24px;display:flex;flex-direction:column;gap:10px}
.report-card{background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);padding:14px 20px;display:flex;align-items:center;justify-content:space-between;gap:16px}
.report-card .date-col{min-width:110px;font-weight:700;color:#1b3a6b}
.report-card .meta{color:#555;font-size:.88rem;flex:1}
.report-card .dl-btn{background:#1b3a6b;color:#fff;border:none;padding:7px 16px;border-radius:6px;cursor:pointer;font-size:.85rem;font-weight:600;white-space:nowrap;text-decoration:none}
.report-card .dl-btn:hover{background:#2a5298}
.report-card .load-btn{background:#e8f0fe;color:#1b3a6b;border:1px solid #c5d5f8;padding:7px 16px;border-radius:6px;cursor:pointer;font-size:.85rem;font-weight:600;white-space:nowrap}
.report-card .load-btn:hover{background:#d2e3fc}

/* Settings */
.settings-wrap{padding:24px;max-width:860px}
.settings-header{margin-bottom:18px}
.settings-header h2{font-size:1.2rem;color:#1b3a6b;margin-bottom:4px}
.settings-header p{color:#666;font-size:.88rem}
.settings-header code{background:#f0f4f8;padding:2px 6px;border-radius:4px;font-family:monospace;font-size:.82rem}
.settings-section{background:#fff;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.1);padding:18px 22px;margin-bottom:14px}
.settings-section h3{font-size:.98rem;color:#1b3a6b;margin-bottom:4px}
.settings-section .hint{color:#777;font-size:.8rem;margin-bottom:10px}
.chip-container{display:flex;flex-wrap:wrap;gap:6px;min-height:32px;padding:4px 0;margin-bottom:8px}
.chip{background:#e8f0fe;color:#1b3a6b;border-radius:16px;padding:4px 10px 4px 12px;font-size:.82rem;display:inline-flex;align-items:center;gap:5px}
.chip button{background:none;border:none;cursor:pointer;font-size:1rem;color:#5c7ab8;line-height:1;padding:0 1px}
.chip button:hover{color:#c62828}
.chip-add-row{display:flex;gap:8px;margin-top:2px}
.chip-add-row input{padding:6px 10px;border:1px solid #c5cdd8;border-radius:6px;font-size:.88rem;flex:1;max-width:300px}
.chip-add-row button{background:#1b3a6b;color:#fff;border:none;padding:6px 16px;border-radius:6px;cursor:pointer;font-size:.85rem;font-weight:600}
.chip-add-row button:hover{background:#2a5298}
.budget-row{display:flex;gap:20px;align-items:flex-end;flex-wrap:wrap}
.budget-row label{font-size:.87rem;color:#444;display:flex;flex-direction:column;gap:5px}
.budget-row input{padding:7px 10px;border:1px solid #c5cdd8;border-radius:6px;font-size:.9rem;width:180px}
.score-row{display:flex;align-items:center;gap:14px;margin-top:6px}
#score-slider{width:220px;accent-color:#1b3a6b;cursor:pointer}
#score-val{font-size:1.3rem;font-weight:700;color:#1b3a6b;min-width:36px}
.time-input{padding:7px 10px;border:1px solid #c5cdd8;border-radius:6px;font-size:.9rem;width:120px}
.settings-footer{margin-top:22px;display:flex;align-items:center;gap:14px}
.save-btn{background:#1a7a3c;color:#fff;border:none;padding:11px 28px;border-radius:6px;font-size:.95rem;font-weight:700;cursor:pointer}
.save-btn:hover{background:#155d2d}
#save-msg{color:#1a7a3c;font-weight:600;font-size:.92rem}
#save-err{color:#c62828;font-weight:600;font-size:.92rem}

/* Scrape overlay */
#scrape-overlay{
  position:fixed;top:0;left:0;width:100%;height:100%;
  background:rgba(10,20,50,.65);
  display:none;align-items:center;justify-content:center;
  z-index:9999;backdrop-filter:blur(3px);
}
#scrape-overlay.visible{display:flex}
.scrape-modal{
  background:#fff;border-radius:16px;padding:40px 48px;
  width:520px;max-width:94%;
  box-shadow:0 28px 72px rgba(0,0,0,.4);
  text-align:center;
}
.scrape-modal-title{font-size:1.3rem;font-weight:700;color:#1b3a6b;margin-bottom:8px}
.scrape-modal-stage{
  font-size:.9rem;color:#555;min-height:2.4em;margin-bottom:22px;
  line-height:1.5;padding:0 8px;
}
.scrape-prog-track{
  background:#dde6f0;border-radius:10px;height:16px;
  overflow:hidden;margin-bottom:10px;
}
.scrape-prog-fill{
  height:100%;width:0%;border-radius:10px;
  background:linear-gradient(90deg,#1b3a6b 0%,#3b82f6 60%,#60a5fa 100%);
  background-size:200% 100%;
  transition:width 1.3s cubic-bezier(.4,0,.2,1);
}
.scrape-prog-fill.animating{
  animation:shimmer 2.2s linear infinite;
}
@keyframes shimmer{
  0%{background-position:200% 0}100%{background-position:-200% 0}
}
.scrape-pct{font-size:2rem;font-weight:800;color:#1b3a6b;margin-bottom:14px;letter-spacing:-1px}
.scrape-modal-hint{font-size:.78rem;color:#aaa;margin-top:6px}
</style>
</head>
<body>

<header>
  <h1>&#128196; Tender Dashboard &mdash; Visakhapatnam</h1>
  <div class="hdr-right">
    <span id="last-run">Last run: &mdash;</span>
    <span id="spinner">&#9203; Scraping&hellip; (~2 min)</span>
    <button id="run-btn" onclick="runScraper()">&#9654; Run Scraper</button>
  </div>
</header>

<div class="tabs">
  <button class="tab active" id="tab-matched" onclick="switchTab('matched')">Matched Tenders</button>
  <button class="tab" id="tab-all" onclick="switchTab('all')">All Tenders (Raw)</button>
  <button class="tab" id="tab-history" onclick="switchTab('history')">&#128190; Past Reports</button>
  <button class="tab" id="tab-settings" onclick="switchTab('settings')">&#9881; Settings</button>
</div>

<div class="controls">
  <input id="search-box" type="text" placeholder="Search title, department, keywords&hellip;" oninput="filterTable()">
  <select id="portal-filter" onchange="filterTable()">
    <option value="">All portals</option>
    <option value="gem">GeM</option>
    <option value="cppp">CPPP (Central eProcure)</option>
  </select>
  <select id="score-filter" onchange="filterTable()">
    <option value="0">All scores</option>
    <option value="80">80+ (Excellent)</option>
    <option value="60">60+ (Good)</option>
    <option value="30">30+ (Possible)</option>
  </select>
  <span id="result-count"></span>
</div>

<div id="tender-view">
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>Score</th>
          <th>Title</th>
          <th class="dept">Department</th>
          <th>Deadline</th>
          <th>Days Left</th>
          <th>Portal</th>
          <th class="kw">Matched Keywords</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
</div>

<div id="history-view" style="display:none">
  <div class="reports-grid" id="reports-grid">
    <div class="empty">Loading past reports&hellip;</div>
  </div>
</div>

<div id="settings-view" style="display:none">
  <div class="settings-wrap">
    <div class="settings-header">
      <h2>&#9881; Search Settings</h2>
      <p>Edit your search criteria below and click <strong>Save Settings</strong>. Changes are written to <code>my_profile.yaml</code> and take effect on the next scrape run.</p>
    </div>

    <div class="settings-section">
      <h3>&#128205; Locations</h3>
      <p class="hint">Tenders mentioning these places get a score bonus. Click &times; to remove, type and press Enter or click + Add to add.</p>
      <div class="chip-container" id="chips-locations"></div>
      <div class="chip-add-row">
        <input type="text" id="add-location" placeholder="e.g. Hyderabad"
               onkeydown="if(event.key==='Enter'){event.preventDefault();addChip('locations','add-location')}">
        <button onclick="addChip('locations','add-location')">+ Add</button>
      </div>
    </div>

    <div class="settings-section">
      <h3>&#128296; Work Type Keywords</h3>
      <p class="hint">A tender must match at least one keyword to appear in your results. Be specific for better precision.</p>
      <div class="chip-container" id="chips-work"></div>
      <div class="chip-add-row">
        <input type="text" id="add-work" placeholder="e.g. roofing"
               onkeydown="if(event.key==='Enter'){event.preventDefault();addChip('work','add-work')}">
        <button onclick="addChip('work','add-work')">+ Add</button>
      </div>
    </div>

    <div class="settings-section">
      <h3>&#128683; Exclude Keywords</h3>
      <p class="hint">Tenders containing any of these words are removed completely from results.</p>
      <div class="chip-container" id="chips-exclude"></div>
      <div class="chip-add-row">
        <input type="text" id="add-exclude" placeholder="e.g. underwater"
               onkeydown="if(event.key==='Enter'){event.preventDefault();addChip('exclude','add-exclude')}">
        <button onclick="addChip('exclude','add-exclude')">+ Add</button>
      </div>
    </div>

    <div class="settings-section">
      <h3>&#128181; Budget Range</h3>
      <p class="hint">Only tenders within this range are considered. Enter amounts in plain rupees (no commas).</p>
      <div class="budget-row">
        <label>Minimum (&#8377;)<input type="number" id="budget-min" step="10000" min="0" placeholder="100000"></label>
        <label>Maximum (&#8377;)<input type="number" id="budget-max" step="100000" min="0" placeholder="20000000"></label>
      </div>
    </div>

    <div class="settings-section">
      <h3>&#127919; Minimum Match Score</h3>
      <p class="hint">Only tenders scoring at least this much appear as &ldquo;Matched&rdquo;. &nbsp;30 = more results &nbsp;|&nbsp; 60 = quality matches &nbsp;|&nbsp; 80 = very precise.</p>
      <div class="score-row">
        <input type="range" id="score-slider" min="10" max="90" step="5"
               oninput="document.getElementById('score-val').textContent=this.value">
        <span id="score-val">30</span>
        <span style="color:#666">/100</span>
      </div>
    </div>

    <div class="settings-section">
      <h3>&#9201; Daily Run Time</h3>
      <p class="hint">What time the scraper runs automatically each day (24-hour format, IST).</p>
      <input type="time" id="run-time" class="time-input">
    </div>

    <div class="settings-footer">
      <button class="save-btn" onclick="saveSettings()">&#128190; Save Settings</button>
      <span id="save-msg" style="display:none">&#10003; Saved successfully!</span>
      <span id="save-err" style="display:none"></span>
    </div>
  </div>
</div>

<!-- Scrape progress overlay — blocks the whole page while running -->
<div id="scrape-overlay">
  <div class="scrape-modal">
    <div class="scrape-modal-title">&#9203;&nbsp; Scraping Tenders&hellip;</div>
    <div class="scrape-modal-stage" id="scrape-stage-msg">Initialising&hellip;</div>
    <div class="scrape-prog-track">
      <div class="scrape-prog-fill animating" id="scrape-prog-fill"></div>
    </div>
    <div class="scrape-pct" id="scrape-pct-label">0%</div>
    <div class="scrape-modal-hint">Please wait &mdash; this usually takes 1&ndash;3 minutes.<br>The page will unlock automatically when the scrape is complete.</div>
  </div>
</div>

<script>
let tab = 'matched';
let matchedData = [], allData = [];
let pollTimer = null;

// ── Startup ────────────────────────────────────────────────────────────────
window.onload = () => { loadData(); checkStatus(); };

// ── Tab ────────────────────────────────────────────────────────────────────
function switchTab(t) {
  tab = t;
  ['matched','all','history','settings'].forEach(id => {
    document.getElementById('tab-' + id).classList.toggle('active', t === id);
  });
  const isTender   = t === 'matched' || t === 'all';
  const isHistory  = t === 'history';
  const isSettings = t === 'settings';
  document.getElementById('tender-view').style.display   = isTender   ? '' : 'none';
  document.getElementById('history-view').style.display  = isHistory  ? '' : 'none';
  document.getElementById('settings-view').style.display = isSettings ? '' : 'none';
  document.querySelector('.controls').style.display      = isTender   ? '' : 'none';
  if (isHistory)  loadReports();
  else if (isSettings) loadSettings();
  else filterTable();
}

// ── Data ───────────────────────────────────────────────────────────────────
async function loadData() {
  const [r1, r2] = await Promise.all([fetch('/api/tenders'), fetch('/api/all-tenders')]);
  matchedData = await r1.json();
  allData     = await r2.json();
  filterTable();
}

// ── Table render ───────────────────────────────────────────────────────────
function scoreClass(s) {
  return s >= 80 ? 'excellent' : s >= 60 ? 'good' : s >= 30 ? 'possible' : 'poor';
}

function filterTable() {
  const q     = (document.getElementById('search-box').value || '').toLowerCase();
  const port  = (document.getElementById('portal-filter').value || '').toLowerCase();
  const minS  = parseInt(document.getElementById('score-filter').value || '0');
  const data  = tab === 'matched' ? matchedData : allData;

  const rows = data.filter(t => {
    if (t.score < minS) return false;
    if (port && !(t.portal||'').toLowerCase().includes(port)) return false;
    if (q && !((t.title||'')+(t.department||'')+(t.keywords||'')).toLowerCase().includes(q)) return false;
    return true;
  });

  document.getElementById('result-count').textContent =
    rows.length + ' tender' + (rows.length !== 1 ? 's' : '');

  const sc = scoreClass;
  document.getElementById('tbody').innerHTML = rows.length === 0
    ? '<tr><td colspan="8" class="empty">No tenders match your filter.</td></tr>'
    : rows.map((t, i) => {
        const cls = sc(t.score);
        const title = t.title ? `<a href="${t.url||'#'}" target="_blank">${esc(t.title)}</a>` : '&mdash;';
        return `<tr class="s-${cls}">
          <td>${i+1}</td>
          <td><span class="badge ${cls}">${t.score}</span></td>
          <td>${title}</td>
          <td class="dept">${esc(t.department||'')}</td>
          <td style="white-space:nowrap">${esc(t.deadline||'\u2014')}</td>
          <td style="white-space:nowrap">${daysLeftBadge(t.days_left)}</td>
          <td style="white-space:nowrap">${esc(t.portal||'')}</td>
          <td class="kw">${esc(t.keywords||'')}</td>
        </tr>`;
      }).join('');
}

function esc(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

function daysLeftBadge(d) {
  if (d === null || d === undefined) return '<span style="color:#aaa">\u2014</span>';
  if (d < 0)  return '<span class="dl-badge dl-expired">Expired</span>';
  if (d === 0) return '<span class="dl-badge dl-urgent">Today!</span>';
  if (d <= 3)  return `<span class="dl-badge dl-urgent">${d}d left</span>`;
  if (d <= 7)  return `<span class="dl-badge dl-soon">${d}d left</span>`;
  return `<span class="dl-badge dl-ok">${d}d left</span>`;
}

// ── Past reports ────────────────────────────────────────────────────────────
async function loadReports() {
  const resp = await fetch('/api/reports');
  const reports = await resp.json();
  const grid = document.getElementById('reports-grid');
  if (!reports.length) {
    grid.innerHTML = '<div class="empty">No past reports found. Run the scraper to create one.</div>';
    return;
  }
  grid.innerHTML = reports.map(r => `
    <div class="report-card">
      <div class="date-col">&#128197; ${r.date}</div>
      <div class="meta">
        <strong>${r.matched}</strong> matched tenders &nbsp;&bull;&nbsp;
        Run at ${r.time} &nbsp;&bull;&nbsp;
        ${r.size_kb} KB
        <br><small style="color:#999">${r.filename}</small>
      </div>
      <a class="dl-btn" href="/reports/${encodeURIComponent(r.filename)}" download>&#11123; Download Excel</a>
      <button class="load-btn" onclick="loadFromReport('${r.filename}')">&#128065; View in Dashboard</button>
    </div>`).join('');
}

async function loadFromReport(filename) {
  const resp = await fetch('/api/report-data/' + encodeURIComponent(filename));
  if (!resp.ok) { alert('Could not load ' + filename); return; }
  const d = await resp.json();
  matchedData = d.matched;
  allData = d.all;
  switchTab('matched');
  document.getElementById('last-run').textContent = 'Viewing: ' + filename.replace('tenders_','').replace('.xlsx','');
}

// ── Scrape overlay helpers ──────────────────────────────────────────────────
function showScrapeOverlay(progress, stage) {
  const overlay = document.getElementById('scrape-overlay');
  overlay.classList.add('visible');
  document.getElementById('scrape-stage-msg').textContent = stage || 'Working\u2026';
  document.getElementById('scrape-pct-label').textContent = (progress || 0) + '%';
  const fill = document.getElementById('scrape-prog-fill');
  fill.style.width = (progress || 0) + '%';
  // Keep shimmer animation while in progress; stop it at 100%
  fill.classList.toggle('animating', progress < 100);
}

function hideScrapeOverlay() {
  document.getElementById('scrape-overlay').classList.remove('visible');
}

// ── Run scraper ────────────────────────────────────────────────────────────
async function runScraper() {
  const resp = await fetch('/run-scraper', {method:'POST'});
  if (resp.status === 409) { alert('Scraper is already running. Please wait.'); return; }
  document.getElementById('run-btn').disabled = true;
  document.getElementById('spinner').style.display = 'inline';
  showScrapeOverlay(0, 'Initialising scrapers\u2026');
  if (!pollTimer) pollTimer = setInterval(checkStatus, 1500);
}

// ── Status polling ─────────────────────────────────────────────────────────
async function checkStatus() {
  const resp = await fetch('/api/status');
  const d = await resp.json();
  document.getElementById('last-run').textContent =
    d.last_run ? 'Last run: ' + d.last_run : 'Last run: \u2014';

  if (d.status === 'running') {
    showScrapeOverlay(d.progress || 0, d.stage || 'Working\u2026');
    document.getElementById('run-btn').disabled = true;
    document.getElementById('spinner').style.display = 'inline';
    // Start polling if the page was opened mid-scrape
    if (!pollTimer) pollTimer = setInterval(checkStatus, 1500);
  } else {
    hideScrapeOverlay();
    document.getElementById('run-btn').disabled = false;
    document.getElementById('spinner').style.display = 'none';
    if (pollTimer) { clearInterval(pollTimer); pollTimer = null; }
    if (d.status === 'done')  loadData();
    if (d.status === 'error') alert('Scraper error: ' + d.error);
  }
}

// ── Settings ────────────────────────────────────────────────────────────────
let _settingsCache = {};

async function loadSettings() {
  const resp = await fetch('/api/settings');
  _settingsCache = await resp.json();

  renderChips('chips-locations', _settingsCache.locations || []);
  renderChips('chips-work',      _settingsCache.my_work_types || []);
  renderChips('chips-exclude',   _settingsCache.exclude_these_work_types || []);

  document.getElementById('budget-min').value   = _settingsCache.budget_minimum || 100000;
  document.getElementById('budget-max').value   = _settingsCache.budget_maximum || 20000000;

  const score = _settingsCache.minimum_match_score || 30;
  document.getElementById('score-slider').value = score;
  document.getElementById('score-val').textContent = score;

  document.getElementById('run-time').value = _settingsCache.run_every_day_at || '08:00';

  document.getElementById('save-msg').style.display = 'none';
  document.getElementById('save-err').style.display = 'none';
}

function renderChips(containerId, items) {
  document.getElementById(containerId).innerHTML = items.map(item =>
    `<span class="chip">${esc(item)}<button type="button" title="Remove" onclick="this.parentElement.remove()">&times;</button></span>`
  ).join('');
}

function addChip(group, inputId) {
  const input = document.getElementById(inputId);
  const val = input.value.trim();
  if (!val) return;
  const containerId = 'chips-' + group;
  const chip = document.createElement('span');
  chip.className = 'chip';
  chip.innerHTML = `${esc(val)}<button type="button" title="Remove" onclick="this.parentElement.remove()">&times;</button>`;
  document.getElementById(containerId).appendChild(chip);
  input.value = '';
  input.focus();
}

function getChips(containerId) {
  return [...document.querySelectorAll('#' + containerId + ' .chip')].map(c =>
    c.firstChild.textContent.trim()
  ).filter(Boolean);
}

async function saveSettings() {
  const saveMsg = document.getElementById('save-msg');
  const saveErr = document.getElementById('save-err');
  saveMsg.style.display = 'none';
  saveErr.style.display = 'none';

  const body = {
    your_name:                _settingsCache.your_name || 'Contractor',
    locations:                getChips('chips-locations'),
    my_work_types:            getChips('chips-work'),
    exclude_these_work_types: getChips('chips-exclude'),
    budget_minimum:           parseInt(document.getElementById('budget-min').value)   || 100000,
    budget_maximum:           parseInt(document.getElementById('budget-max').value)   || 20000000,
    minimum_match_score:      parseInt(document.getElementById('score-slider').value) || 30,
    run_every_day_at:         document.getElementById('run-time').value || '08:00',
  };

  try {
    const resp = await fetch('/api/settings', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(body),
    });
    const d = await resp.json();
    if (resp.ok) {
      _settingsCache = Object.assign(_settingsCache, body);
      saveMsg.style.display = 'inline';
      setTimeout(() => { saveMsg.style.display = 'none'; }, 3000);
    } else {
      saveErr.textContent = 'Error: ' + (d.error || 'unknown');
      saveErr.style.display = 'inline';
    }
  } catch (e) {
    saveErr.textContent = 'Network error: ' + e.message;
    saveErr.style.display = 'inline';
  }
}
</script>
</body>
</html>"""

# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    _load_latest_excel()
    print()
    print("  Tender Dashboard is running.")
    print("  Open this link in your browser:  http://localhost:5001")
    print("  Press Ctrl+C to stop.")
    print()
    app.run(debug=False, host="127.0.0.1", port=5001)
